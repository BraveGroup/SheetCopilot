import xlwings as xw
from xlwings import Range
from xlwings import constants as win32c
from win32com.client import constants as c
from  .constants import constants
import win32api
import win32com.client as win32
from .AtomicAction import Action
from functools import wraps
from typing import Any, Tuple, Callable, Optional, Union, List
from openpyxl.utils import get_column_letter
import sys, os, re
import itertools

SUMMARIZATION_DICT = {
    'sum': c.xlSum 
    # 'sum', 'count', 'average', 'max', 'min', 'product', 'countNumbers', 'standardDeviation', 'standardDeviationP', 'var', or 'varP'
}
class xwBackend():
    '''
    Concrete Action implement the atomic actions by xlwings.
    '''

    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = True
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRange(self, source: str):
        '''
        Converts a string to a Range object.
        
        Parameters:
        source (str): The string to be converted to a Range object.
        
        Returns:
        Range: The Range object.
        '''
        # sheet = self.activeWS
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") # Sheet names with spaces are enclosed with single quotes which should be removed
            # check if the sheet exists
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        '''
        Opens a workbook.

        Parameters:
        path (str): The path to the workbook.

        Returns:
        None
        '''
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        '''
        Saves a workbook.

        Parameters:
        path (str): The path to the workbook.

        Returns:
        None
        '''
        self.activeWB.SaveAs(os.path.abspath(path))
                
    def Write(self, range: str, value: Any) -> None:
        '''
        Writes a value to a Range.

        Parameters:
        range (str): The Range object to write to.
        value (Any): The value to write to the Range.

        Returns:
        None
        '''
        range = self.toRange(range)
        
        if range.Rows.Count > 5000:
            raise Exception('The range is too large. Please note that reference to whole columns like A:A and C:D actually includes millions of rows. Please use a specific range like A1:C30 instead.')
        
        if isinstance(value, (list, tuple)) and range.Count == 1:
            if isinstance(value[0], (list, tuple)):
                for rowOffet, elem in enumerate(value):
                    for columnOffset, elem2 in enumerate(elem):
                        range.GetOffset(rowOffet, columnOffset).Value = elem2
            else:
                for columnOffset, elem in enumerate(value):
                    range.GetOffset(0, columnOffset).Value = elem
        else:
            range.Value = value

    def CopyPaste(self, source: str, destination: str) -> None:
        '''
        Copies the value and format of a source Range to a destination Range.

        Parameters:
        source (str): The Range object to copy from.
        destination (str): The Range object to copy to.

        Returns:
        None
        '''
        source = self.toRange(source).SpecialCells(12)
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4163)
        # self.CopyPasteVisible(source, destination)

    def CopyPasteVisible(self, source: str, destination: str) -> None:
        '''
        Copies the visible value and format of a source Range to a destination Range.

        Parameters:
        source (str): The Range object to copy from.
        destination (str): The Range object to copy to.

        Returns:
        None
        '''
        source = self.toRange(source)
        rowCount, columnCount = source.Rows.Count, source.Columns.Count
        firstCell = source.Cells(1,1)
        while firstCell.EntireRow.Hidden:
            firstCell = firstCell.GetOffset(1,0)
        while firstCell.EntireColumn.Hidden:
            firstCell = firstCell.GetOffset(0,1)
        lastCell = firstCell
        while rowCount > 0:
            lastCell = lastCell.GetOffset(1,0)
            if not lastCell.EntireRow.Hidden:
                rowCount -= 1
        while columnCount > 0:
            lastCell = lastCell.GetOffset(0,1)
            if not lastCell.EntireColumn.Hidden:
                columnCount -= 1
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4163)
    
    def CopyPasteFormat(self, source: str, destination: str) -> None:
        '''
        Copies the format of a source Range to a destination Range.

        Parameters:
        source (Range): The Range object to copy format from.
        destination (Range): The Range object to copy format to.

        Returns:
        None
        '''
        source = self.toRange(source)
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4122)

    def CutPaste(self, source: str, destination: str) -> None:
        '''
        Cuts the value and format of a source Range to a destination Range.

        Parameters:
        source (Range): The Range object to cut from.
        destination (Range): The Range object to cut to.

        Returns:
        None
        '''
        source = self.toRange(source)
        destination = self.toRange(destination)
        source.Cut(Destination=destination)

    def FindReplace(self, source: str, find: str, replace: str) -> None:
        '''
        Finds and replaces the value of a source Range.

        Parameters:
        source (str): The Range object to find and replace.
        old (str): The value to find.
        new (str): The value to replace.

        Returns:
        None
        '''
        source = self.toRange(source)
        source.Replace(find, replace)

    def SetHyperlink(self, source: str, url: str) -> None:
        '''
        Sets a hyperlink to a Range.

        Parameters:
        source (str): The Range object to set the hyperlink to.
        link (str): The hyperlink to set.

        Returns:
        None
        '''
        source = self.toRange(source)
        sheet = source.Parent
        sheet.Hyperlinks.Add(Anchor=source, Address=url, TextToDisplay=str(source.Value))

    def RemoveHyperlink(self, source: str) -> None:
        '''
        Removes the hyperlink of a Range.

        Parameters:
        source (str): The Range object to remove the hyperlink from.
        
        Returns:
        None
        '''
        source = self.toRange(source)
        source.ClearHyperlinks()

    def RenameSheet(self, oldName: str, newName: str) -> None:
        self.activeWB.Sheets(oldName).Name = newName

    def WrapText(self, range: str) -> None:
        range = self.toRange(range)
        range.WrapText = True

    def UnwrapText(self, range: str) -> None:
        range = self.toRange(range)
        range.api.WrapText = False

    def AutoFill(self, source: str, destination: str) -> None:
        '''
        Autofills the value of a source Range to a destination Range.

        Parameters:
        source (str): The Range object to autofill from.
        destination (str): The Range object to autofill to.

        Returns:
        None
        '''
        source = self.toRange(source)
        destination = self.toRange(destination)
        
        # source_row_start, source_row_end, source_column_start, source_column_end = source.Row, source.Row + source.Rows.Count, source.Column, source.Column + source.Columns.Count
        # destination_row_start, destination_row_end, destination_column_start, destination_column_end = destination.Row, destination.Row + destination.Rows.Count, destination.Column, destination.Column + destination.Columns.Count
        
        # AutoFill can be done vertically or horizontally.
        # Check if the destination range includes the source.
        
        if self.activeAPP.Union(destination, source).Address != destination.Address:
            raise ValueError('Illegal source and destination! The auto-filling destination must include the source!')

        source.AutoFill(destination)

    def Sort(self, source: str, key1: str, order: str='asc', orientation: str='column') -> None:
        '''
        Sort the source Range based on the key1 Range.

        Parameters:
        source (str): The Range object to sort.
        key1 (str): The Range is used as the key for sorting.
        order (str): one of 'asc' and 'dec'.
        orientation (str): one of 'column' and 'row'.

        Returns:
        None
        '''
        source = self.toRange(source)
        key1 = self.toRange(key1)
        source.Sort(Key1=key1, Order1=1 if order == 'asc' else 2, 
                        Orientation=1 if orientation == 'column' else 2)

    def Filter(self, source: str, fieldIndex: int, criteria: str) -> None:
        '''
        Filter the source Range based on the key1 Range.

        Parameters:
        source (str): The Range object to filter.
        field (int): The integer offset of the field on which you want to base the filter, from the left of the list, the leftmost field is field one.
        criteria (str): The criteria used to filter source Range.

        Returns:
        None
        '''
        source = self.toRange(source)
        try:
            criteriaRange = self.toRange(criteria)
        except:
            criteriaRange = None
        if criteriaRange:
            criteria = [criteriaRange.Cells(i).Text for i in range(1, criteriaRange.Cells.Count + 1)]
        source.AutoFilter(Field=fieldIndex, Criteria1=criteria)
        # source.AutoFilter(Field=field, Criteria1=criteria, Operator=constants.AutoFilterOperator['values'])

    def DeleteFilter(self) -> None:
        """
        Delete all filters.

        Parameters:
        None.

        Returns:
        None
        """
        if self.activeWS.AutoFilterMode: self.activeWS.AutoFilterMode = False

    def MoveRow(self, source: int, destination: int) -> None:
        """
        Move a row from one position to another.
        
        Parameters:
        source (int): The row number to move.
        destination (int): The row number to move to.
        
        Returns:
        None
        """
        if destination > source:
            destination += 1
        else:
            source += 1
        self.InsertRow(destination, aboveRow=destination)
        self.activeWS.Rows(source).Copy(self.activeWS.Rows(destination))
        self.activeWS.Rows(source).Delete()

    def MoveColumn(self, source: int, destination: int) -> None:
        """
        Move a column from one position to another.

        Parameters:
        source (int): The column number to move.
        destination (int): The column number to move to.

        Returns:
        None
        """
        if destination > source:
            destination += 1
        else:
            source += 1
        self.InsertColumn(destination)
        self.activeWS.Columns(source).Copy(self.activeWS.Columns(destination))
        self.activeWS.Columns(source).Delete()
    
    def RemoveDuplicate(self, source: str, key: int) -> None:
        """
        Removes duplicate values from a range of values.

        Parameters:
        source (str): The Range object to process.
        key (int): The integer offset of the field on which you want to remove duplicate, from the left of the list, the leftmost field is field one.

        Returns:
        None
        """
        source = self.toRange(source)
        source.RemoveDuplicates(key)

    def group_ungroup(self) -> None:
        pass

    def SetPassword(self, password: str) -> None:
        """
        Sets the password of the active workbook.
        
        Parameters:
        password (str): The password to set.
        
        Returns:
        None
        """
        self.activeWB.Password = password

    def TransposeRange(self, source: str) -> None:
        source = self.toRange(source)
        dataT = self.activeAPP.WorksheetFunction.Transpose(source)
        source.Clear()
        cell = source.Cells(1).Address
        self.Write(cell, dataT)

    def CreateNamedRange(self, source: str, name: str):
        source = self.toRange(source)
        source.Name = name

    def SetFormat(self, source: str, font: Optional[str] = None, fontSize: Optional[float] = None,
                    color: Optional[int] = None, fillColor: Optional[int] = None, bold: Optional[bool] = None,
                    italic: Optional[bool] = None, underline: Optional[bool] = None, horizontalAlignment: Optional[str] = None) -> None:
        source = self.toRange(source)
        if font:
            source.Font.Name = font
        if fontSize:
            source.Font.Size = fontSize
        if color:
            source.Font.ColorIndex = constants.ColorIndex[color]
        if fillColor:
            source.Interior.ColorIndex = constants.ColorIndex[fillColor]
        if not bold is None:
            source.Font.Bold = bold
        if not italic is None:
            source.Font.Italic = italic
        if not underline is None:
            source.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone
        if horizontalAlignment:
            source.HorizontalAlignment = constants.HorizontalAlignment[horizontalAlignment]
    
    def DeleteFormat(self, source: str) -> None:
        source = self.toRange(source)
        source.ClearFormats()

    def SetDataType(self, source: str, dataType: str) -> None:
        source = self.toRange(source)
        source.NumberFormat = constants.DataType[dataType]

    def SetPageLayout(self, orientation: str, paperSize: str) -> None:
        '''
        Set the page layout.
        
        Parameters:
        orientation (str): The orientation of the page, 'landscape' or 'portrait'.
        paperSize (str): The size of the page.

        Returns:
        None
        '''
        self.activeWS.PageSetup.Orientation = constants.PageOrientation[orientation]
        self.activeWS.PageSetup.PaperSize = constants.PaperSize[paperSize]

    def SetBorderAround(self, source: str, color: str, weight: str) -> None:
        '''
        Set the border around a range.

        Parameters:
        source (str): The range of cells which you want to set the border.
        color (str): The color of the border.
        weight (str): The weight of the border. hairline, 'thin', 'medium', 'thick'.

        Returns:
        None
        '''
        source = self.toRange(source)
        source.BorderAround(ColorIndex=constants.ColorIndex[color], Weight=constants.BorderWeight[weight])

    def ToggleRowColumnVisibility(self, range: str, visible: bool, region: str) -> None:
        '''
        Toggle the visibility of rows or columns.

        Parameters:
        range (str): The range of rows or columns which you want to toggle the visibility.
        visible (bool): True to show the rows or columns, False to hide them.
        region (str): 'row' or 'column'.

        Returns:
        None
        '''
        range = self.toRange(range)
        if region == 'row':
            range.EntireRow.Hidden = not visible
        elif region == 'column':
            range.EntireColumn.Hidden = not visible

    def SetCellMerge(self, source: str, merge: bool) -> None:
        '''
        Toggle the merge of cells.

        Parameters:
        range (str): The range of cells which you want to toggle the merge.
        merge (bool): True to merge the cells, False to unmerge them.

        Returns:
        None
        '''
        source = self.toRange(source)
        if merge:
            source.Merge()
        else:
            source.Unmerge()

    def merging_text(self) -> None:
        pass

    def Delete(self, source: str, region: str) -> None:
        """
        Deletes a cell or range of cells.

        Parameters:
        source (str): The Range object to delete.

        Returns:
        None
        """
        source = self.toRange(source)
        if region == 'row':
            source.EntireRow.Delete()
        elif region == 'column':
            source.EntireColumn.Delete()
        else:
            source.Delete()

    def Clear(self, source: str) -> None:
        """
        Clears the content and the formatting of a Range.

        Parameters:
        source (Range): The Range object to clear.

        Returns:
        None
        """
        source = self.toRange(source)
        source.Clear()

    def Insert(self, source: str, shift: Optional[str]) -> None:
        """
        Insert a cell or range of cells into the sheet.

        Parameters:
        source (str): The Range object to indicate the postion to insert.
        shift (str, default None):  Use 'right' or 'down'. If omitted, Excel decides based on the shape of the range.

        Returns:
        None
        """
        source = self.toRange(source)
        source.Insert(constants.InsertShiftDirection[shift])

    def InsertRow(self, sheetName: str, aboveRow: int = None, belowRow: int = None) -> None:
        """
        Insert one row.

        Parameters:
        sheetName (str): The name of the sheet.
        aboveRow (int): The row above which the new row will be inserted.
        belowRow (int): The row below which the new row will be inserted.

        Returns:
        None
        """
        if aboveRow:
            index = aboveRow
        elif belowRow:
            index = belowRow + 1
        sheet = self.activeWB.Sheets(sheetName)
        lastCell = sheet.UsedRange(sheet.UsedRange.Count)
        source = sheet.Range(sheet.Cells(index, 1), sheet.Cells(index, lastCell.Column))
        source.Insert(constants.InsertShiftDirection['down'])
    
    def InsertColumn(self, sheetName: str, beforeColumn: str = None, afterColumn: str = None) -> None:
        """
        Insert one column.

        Parameters:
        sheetName (str): The name of the sheet.
        beforeColumn (str): The column before which the new column will be inserted.
        afterColumn (str): The column after which the new column will be inserted.

        Returns:
        None
        """
        if beforeColumn:
            index = self.activeWS.Columns(beforeColumn).Column
        elif afterColumn:
            index = self.activeWS.Columns(afterColumn).Column + 1
        sheet = self.activeWB.Sheets(sheetName)
        lastCell = sheet.UsedRange(sheet.UsedRange.Count)
        source = sheet.Range(sheet.Cells(1, index), sheet.Cells(lastCell.Row, index))
        source.Insert(constants.InsertShiftDirection['right'])

    def SplitText(self, source: str, delimiter: str) -> None:
        """
        Splite text in one column to multiple column.

        Parameters:
        source (str): The one column range.
        delimiter (str): The delimiter.

        Returns:
        None
        """
        source = self.toRange(source)
        if source.Columns.Count != 1:
            print('Source must be one column.')
            return
        row = source.Row
        column = source.Column
        orgData = source.Value
        newData = [x.split(delimiter) for x in orgData]
        maxLen = max(len(x) for x in newData)
        newData = [x + [None] * (maxLen-len(x)) for x in newData]
        for i in range(maxLen - 1):
            self.InsertColumn(source.Column)
        self.Write(self.activeWS.Cells(row,column).Address, newData)

    def AutoFit(self, source: str) -> None:
        """
        Autofits the width and height of all cells in the range.

        Parameters:
        source (str): The Range object to process.

        Returns:
        None
        """
        source = self.toRange(source)
        source.AutoFit()

    def ResizeRowColumn(self, source: str, height: Optional[int] = None, width: Optional[int] = None) -> None:
        """
        Resizes the height and width of a range of cells.

        Parameters:
        source (str): The Range object to process.
        height (int): The height of the range.
        width (int): The width of the range.

        Returns:
        None
        """
        source = self.toRange(source)
        if height:
            source.RowHeight = height
        if width:
            source.ColumnWidth = width

    def SetConditionalFormat(self, source: str, formula: str,
                            bold: Optional[bool] = None, color: Optional[str] = None,
                            fillColor: Optional[str] = None, italic: Optional[bool] = None, underline: Optional[bool] = None) -> None:
        """
        Applies conditional formatting to a range of cells.

        Parameters:
        source (str): The Range object to process.
        type (str): The type of conditional format.
        operator (str): The operator of conditional format.
        Formula1 (str): The value or expression associated with the conditional format. Can be a constant value, a string value, a cell reference, or a formula.
        bold (bool, default None): Whether to bold the text.
        color (str, default None): The color of the text.
        fillColor (str, default None): The color of the background.
        italic (bool, default None): Whether to italic the text.
        underline (bool, default None): Whether to underline the text.

        Returns:
        None
        """
        source = self.toRange(source)

        handle = source.FormatConditions.Add(Type = constants.FormatConditionType['expression'], Formula1 = formula)
        if color:
            handle.Font.ColorIndex = constants.ColorIndex[color]
        if fillColor:
            handle.Interior.ColorIndex = constants.ColorIndex[fillColor]
        if not bold is None:
            handle.Font.Bold = bold
        if not italic is None:
            handle.Font.Italic = italic
        if not underline is None:
            handle.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone

    def SetDataValidation(self, source: str, type: str, formula1: str) -> None:
        source = self.toRange(source)
        handle = source.Validation.Add(constants.ValidationType[type], Formula1 = formula1)

    def ToggleFormulaDisplay(self, display: bool) -> None:
        """
        Display or hide formulas.

        Parameters:
        display (bool): Whether to display formulas.

        Returns:
        None
        """
        self.activeAPP.ActiveWindow.DisplayFormulas = display

    def SplitPanes(self, rowCount: int, columnCount: int) -> None:
        """
        Split panes.

        Parameters:
        rowCount (int): The number of rows.
        columnCount (int): The number of columns.

        Returns:
        None
        """
        self.activeAPP.ActiveWindow.SplitRow = rowCount
        self.activeAPP.ActiveWindow.SplitColumn = columnCount

    def SetCellLock(self, source: str, lock: bool) -> None:
        """
        Lock or unlock cells.

        Parameters:
        source (str): The Range object to process.
        lock (bool): Whether to lock the cells.

        Returns:
        None
        """
        source = self.toRange(source)
        source.Locked = lock

    def ToggleSheetProtection(self, sheetName: str, protect: bool, password: str = None) -> None:
        """
        Protect or unprotect a sheet.

        Parameters:
        sheetName (str): The name of the sheet.
        protect (bool): Whether to protect the sheet.
        password (str, default None): The password to protect the sheet.

        Returns:
        None
        """
        sheet = self.activeWB.Worksheets(sheetName)
        if protect:
            sheet.Protect(password)
        else:
            sheet.Unprotect(password)

    def FreezePanes(self, source: str) -> None:
        """
        Freezes the panes in a window.

        Parameters:
        source (str): all rows above, and all columns left on source will be frozen.

        Returns:
        None
        """
        source = self.toRange(source)
        source.Select()
        source.Parent.Application.ActiveWindow.FreezePanes = True

    def UnfreezePanes(self, sheetName: Optional[str] = None) -> None:
        """
        Unfreezes the panes in a window.

        Parameters:
        sheetName (str, default None): The name of the sheet. If omitted, the active sheet will be used.

        Returns:
        None
        """
        if sheetName:
            sheet = self.activeWB.Worksheets(sheetName)
        else:
            sheet = self.activeWS
        sheet.Application.ActiveWindow.FreezePanes = False

    def CreateChart(self, source: str, destSheet: str, chartType: str, chartName: str, XField: int = None, YField: List[int] = []) -> None:
        # check if the chart name exists
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    raise ValueError(f'The chart name {chartName} already exists.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        dataRange = self.toRange(source)
        destRange = self.GetBlankArea(destSheet)
        sheet = self.activeWB.Worksheets(destSheet)
        chart = sheet.ChartObjects().Add(destRange.Left, destRange.Top, 350, 200).Chart
        chart.ChartType = constants.ChartType[chartType]
        if 'pie' in chartType.lower():
            chart.SetSourceData(dataRange)
        else:
            if not XField:
                XField = 1
            XFieldRange = dataRange.Parent.Range(dataRange.Cells(2, XField), dataRange.Cells(dataRange.Rows.Count, XField))
            if not YField:
                YField = [i for i in range(1, dataRange.Columns.Count + 1) if i != XField]
            for i in YField:
                series = chart.SeriesCollection().NewSeries()
                series.XValues = XFieldRange
                series.Values = dataRange.Parent.Range(dataRange.Cells(2, i), dataRange.Cells(dataRange.Rows.Count, i))
                series.Name = dataRange.Cells(1, i)
            try:
                chart.Axes(constants.AxisType['x']).CategoryNames = XFieldRange
            except:
                pass
        chart.Parent.Name = chartName

    def SetChartTrendline(self, chartName: str, trendlineType: List[str], DisplayEquation: Optional[bool] = None,
                          DisplayRSquared: Optional[bool] = None) -> None:
        """
        Set trendline for chart.

        Parameters:
        name (str): The name of the chart.
        trendlineType (str): The type of the trendline.
        DisplayEquation (bool, default None): Whether to display the equation on the chart.
        DisplayRSquared (bool, default None): Whether to display the R-squared value on the chart.

        Returns:
        None
        """
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        trendlineType = itertools.cycle(trendlineType)
        for series in chart.SeriesCollection():
            for trendline in series.Trendlines():
                trendline.Delete()
            series.Trendlines().Add(
                constants.TrendlineType[next(trendlineType)],
                DisplayEquation=DisplayEquation,
                DisplayRSquared=DisplayRSquared
            )

    def SetChartTitle(self, chartName: str, title: str, fontSize: Optional[float] = None, 
                        bold: bool = None, color: Optional[int] = None) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasTitle = True
        chart.ChartTitle.Text = title
        if fontSize:
            chart.ChartTitle.Font.Size = fontSize
        if color:
            chart.ChartTitle.Font.ColorIndex = {
                'black': 1,
                'white': 2,
                'red': 3,
                'green': 4,
                'blue': 5,
                'yellow': 6,
                'magenta': 7,
                'cyan': 8,
                'dark red': 9,
                'dark green': 10
            }[color]
        if not bold is None:
            chart.ChartTitle.Font.Bold = bold

    def SetChartHasAxis(self, chartName: str, axis: str, hasAxis: bool) -> None:
        '''
        Set whether the chart has axis.
        
        Parameters:
        name (str): The name of the chart.
        axis (str): The axis to set. 'x' or 'y'.
        hasAxis (bool): Whether the chart has axis.

        Returns:
        None
        '''
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        axis = constants.AxisType[axis]
        axisGroup = constants.AxisGroup['primary']
        chart.SetHasAxis(axis, axisGroup, hasAxis)
        

    def SetChartAxis(self, chartName: str, axis: str, title: Optional[str] = None, 
                        labelOrientation: Optional[str] = None, maxValue: Optional[float] = None,
                        miniValue: Optional[float] = None) -> None:
        '''
        Set the chart axis.

        Parameters:
        name (str): The name of the chart.
        axis (str): The axis to set. 'x' or 'y'.
        title (str, default None): The title of the axis.
        labelOrientation (str, default None): The orientation of the label.
        maxValue (float, default None): The max value of the axis.
        miniValue (float, default None): The mini value of the axis.

        Returns:
        None
        '''
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        if axis in ['x', 'X']:
            axis = win32c.AxisType.xlCategory
        elif axis in ['y', 'Y']:
            axis = win32c.AxisType.xlValue
        elif axis in ['z', 'Z']:
            axis = win32c.AxisType.xlSeriesAxis
        else:
            print('Not support axes type')
            return
        chartAxes = chart.Axes(axis)
        if title:
            chartAxes.HasTitle = True
            chartAxes.AxisTitle.Text = title
        if labelOrientation:
            labelOrientation = {
                'upward': win32c.Orientation.xlUpward,
                'downward': win32c.Orientation.xlDownward,
                'horizontal': win32c.Orientation.xlHorizontal,
                'vertical': win32c.Orientation.xlVertical
                }[labelOrientation]
            chartAxes.TickLabels.Orientation = labelOrientation
        if maxValue:
            chartAxes.MaximumScale = maxValue
        if miniValue:
            chartAxes.MinimumScale = miniValue

    def SetChartLegend(self, chartName: str, position: Optional[str] = None, fontSize: Optional[str] = None,
                        seriesName: Optional[list] = []) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = True
        
        if position and position != 'None':
            # For legent position enumeration, refer to https://learn.microsoft.com/en-us/dotnet/api/microsoft.office.interop.word.xllegendposition?view=word-pia
            position = {
                'bottom': win32c.LegendPosition.xlLegendPositionBottom,
                'corner': win32c.LegendPosition.xlLegendPositionCorner,
                'left': win32c.LegendPosition.xlLegendPositionLeft,
                'right': win32c.LegendPosition.xlLegendPositionRight,
                'top': win32c.LegendPosition.xlLegendPositionTop
            }[position]
            chart.Legend.Position = position
        if seriesName:
            for index, elem in enumerate(seriesName):
                chart.SeriesCollection(index+1).Name = elem

    def SetChartHasLegend(self, chartName: str, hasLegend: bool) -> None:
        '''
        Set whether the chart has legend.

        Parameters:
        name (str): The name of the chart.
        hasLegend (bool): Whether the chart has legend.

        Returns:
        None
        '''
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = hasLegend

    def SetChartType(self, chartName: str, chartType: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartType = constants.ChartType[chartType]

    def SetChartSource(self, chartName: str, source: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        source = self.toRange(source)
        chart.SetSourceData(source)

    def SetChartBackgroundColor(self, chartName: str, color: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartArea.Interior.ColorIndex = constants.ColorIndex[color]

    def ResizeChart(self, chartName: str, width: float, height: float) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart.Width = width
        chart.Height = height

    def SetChartDataColor(self, chartName: str, colorRGB: list) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            for point in series.Points():
                point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)
    
    def HighlightDataPoints(self, chartName: str, pointIndex: int, colorRGB: list) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(1)
        point = series.Points(pointIndex)
        point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)

    def SetDataSeriesType(self, chartName: str, seriesIndex: int, seriesType: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.ChartType = constants.ChartType[seriesType]

    def AddDataSeries(self, chartName: str, xrange: str, yrange: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        xrange = self.toRange(xrange)
        yrange = self.toRange(yrange)
        series = chart.SeriesCollection().NewSeries()
        series.XValues = xrange
        series.Values = yrange
    
    def RemoveDataSeries(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.Delete()

    def SetDataSeriesSource(self, chartName: str, seriesIndex: int, xrange: str, yrange: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        xrange = self.toRange(xrange)
        yrange = self.toRange(yrange)
        series = chart.SeriesCollection(seriesIndex)
        series.XValues = xrange
        series.Values = yrange

    def AddChartErrorBars(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = True

    def AddChartErrorBar(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = True

    def RemoveChartErrorBars(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = False

    def RemoveChartErrorBar(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = False

    def AddDataLabels(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = True
    
    def RemoveDataLabels(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = False

    def SetChartMarker(self, chartName: str, style: List[str] = None, size: Optional[float] = None) -> None:
        '''
        style: auto, circle, dash, dot, star, triangle, square, plus
        '''
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        styleIter = itertools.cycle(style)
        for series in chart.SeriesCollection():
            if style:
                series.MarkerStyle = constants.MarkerStyle[next(styleIter)]
            if size:
                series.MarkerSize = size

    def CopyPasteChart(self, chartName, destination: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        destination = self.toRange(destination)
        chart.ChartArea.Copy()
        destination.Select()
        destination.Parent.Paste()

    def CreatePivotTable(self, source: str, destSheet: str, name: str,
                        RowField: List = [], ColumnField: List = [],
                        PageField: List = [], DataField: List = [],
                        summarizeFunction = 'sum') -> None:
        '''
        Create a pivot table.

        source: the source data range
        destSheet: the sheet name to put the pivot table
        name: the name of the pivot table
        RowField: the row fields
        ColumnField: the column fields
        PageField: the page fields
        DataField: the data fields
        summarizeFunction: 'sum', 'count', 'average', 'max', 'min', 'product', 'countNumbers', 'standardDeviation', 'standardDeviationP', 'var', or 'varP'
        
        Example:
        CreatePivotTable('A1:B10', 'PivotTable1', RowField=['Column A'], ColumnField=['Column B'], DataField=['Column B'])
        '''
        # check if the pivot table name exists
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    raise ValueError(f'Pivot table {name} already exists. Please choose a different name.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        # # check if the four fields are letters
        # invalid_fields = []
        # if any([len(x) > 1 for x in RowField]):
        #     invalid_fields.append('RowField')
        # if any([len(x) > 1 for x in ColumnField]):
        #     invalid_fields.append('ColumnField')
        # if any([len(x) > 1 for x in PageField]):
        #     invalid_fields.append('PageField')
        # if any([len(x) > 1 for x in DataField]):
        #     invalid_fields.append('DataField')
        
        # if len(invalid_fields) > 0:
        #     raise ValueError('Illegal fields! the fields in {} can only be column indices (i.e., letters A to Z)'.format(",".join(invalid_fields)))
        
        sheet = self.activeWB.Worksheets(destSheet)
        sourceRange = self.toRange(source)
        
        # Sometimes the LLM misses the header row, so we manually add it
        if sourceRange.Row != 1:
            new_starting_cell = sheet.Cells(1, sourceRange.Column)
            sourceRange = sheet.Range(new_starting_cell, sheet.Cells(new_starting_cell.Row + sourceRange.Rows.Count, sourceRange.Column + sourceRange.Columns.Count - 1))
            
        pc = self.activeWB.PivotCaches().Create(SourceType=win32c.PivotTableSourceType.xlDatabase, SourceData=sourceRange)
        destRange = self.GetBlankArea(destSheet)
        pc.CreatePivotTable(TableDestination=destRange, TableName=name)
        pt = sheet.PivotTables(name)
        for filed in RowField:
            pt.PivotFields(filed).Orientation = win32c.PivotFieldOrientation.xlRowField
        for filed in ColumnField:
            pt.PivotFields(filed).Orientation = win32c.PivotFieldOrientation.xlColumnField
        for filed in PageField:
            pt.PivotFields(filed).Orientation = win32c.PivotFieldOrientation.xlPageField
        for filed in DataField:
            pt.PivotFields(filed).Orientation = win32c.PivotFieldOrientation.xlDataField
            # pt.PivotFields(filed).Function = constants.ConsolidationFunction[summarizeFunction]

    def GetBlankArea(self, sheetName: str):
        '''
        Get the blank area in the worksheet

        Parameters:
        sheetName: the name of the worksheet
        
        Returns:
        Range -- the blank area
        '''
        sheet = self.activeWB.Sheets(sheetName)
        chartRangeList = []
        for chart in sheet.ChartObjects():
            chartRangeList.append(sheet.Range(chart.TopLeftCell, chart.BottomRightCell))
        row, column = 1, 1
        checkScope = 5
        while True:
            cell1 = sheet.Cells(row,column)
            cell2 = sheet.Cells(row+checkScope,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if all(cell.Value is None for cell in checkRange) and all(self.activeAPP.Intersect(chartRange, checkRange) is None for chartRange in chartRangeList):
                break
            row += 1
            column += 1

        while row > 1:
            cell1 = sheet.Cells(row-1,column)
            cell2 = sheet.Cells(row-1,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            row -= 1
            
        while column > 1:
            cell1 = sheet.Cells(row,column-1)
            cell2 = sheet.Cells(row+checkScope,column-1)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            column -= 1

        return sheet.Cells(row+1,column+1)

    def CreateChartFromPivotTable(self, pivotTableName: str, destSheet: str, chartName: str, chartType: str) -> None:
        '''
        Create a pivot chart based on pivot table.

        pivotTableName: the name of the pivot table
        chartName: the name of the chart
        chartType: the type of the chart
        '''
        # find the pivot table
        for sheet in self.activeWB.Worksheets:
            pt_name = None
            for pt in sheet.PivotTables():
                pt_name = pt.Name
                print(pt_name, '|', pivotTableName)
                if pt_name == pivotTableName:
                    break
            if pt_name is not None: break
        else:
            pt = None

        if pt is None:
            raise ValueError(f'Pivot table {pivotTableName} does not exist. Note that this API is only for creating chart from data in pivot table.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        sourceRange = pt.TableRange2
        destRange = self.GetBlankArea(destSheet)
        sheet = self.activeWB.Worksheets(destSheet)
        chart = sheet.ChartObjects().Add(destRange.Left, destRange.Top, 350, 200).Chart
        chart.ChartType = constants.ChartType[chartType]
        chart.SetSourceData(sourceRange)
        chart.Parent.Name = chartName
        
    def RemovePivotTable(self, name: str) -> None:
        # find the pivot table
        pt = None
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    break
        if pt is None or pt.Name != name:
            raise ValueError(f'Pivot table {name} does not exist.')
        pt.TableRange2.Clear()

    def SetPivotTableSummaryFunction(self, name: str, field: str, func: str) -> None:
        pt = self.activeWS.PivotTables(name)
        pt.PivotFields(field).Function = constants.ConsolidationFunction[func]

    def SortPivotTable(self, name: str, field: str, key: str, oder: str = 'ascending') -> None:
        pt = self.activeWS.PivotTables(name)
        oder = constants.SortOrder[oder]
        pt.PivotFields(field).AutoSort(oder, key)

    def CreateSheet(self, sheetName: Optional[str] = None, before: Optional[str] = None, after: Optional[str] = None) -> None:
        '''
        Create a new sheet.

        name: the name of the new sheet
        before: the name of the sheet before which the new sheet created
        after: the name of the sheet after which the new sheet created
        '''
        for sheet in self.activeWB.Worksheets:
            if sheet.Name == sheetName:
                raise ValueError(f'Sheet {sheetName} already exists.')
        activeSheet = self.activeWS
        if before is not None:
            beforeSheet = self.activeWB.Worksheets(before)
            newSheet = self.activeWB.Worksheets.Add(Before=beforeSheet)
        elif after is not None:
            afterSheet = self.activeWB.Worksheets(after)
            newSheet = self.activeWB.Worksheets.Add(After=afterSheet)
        else:
            newSheet = self.activeWB.Worksheets.Add()
        if sheetName is not None:
            newSheet.Name = sheetName
        activeSheet.Activate()

    def RemoveSheet(self, sheetName: Optional[str] = None) -> None:
        '''
        Remove a sheet.

        name: the name of the sheet. If not specified, the active sheet will be removed.
        '''

        if not sheetName:
            sheetName = self.activeWS.Name
        self.activeWB.Sheets(sheetName).Delete()

    def SwitchSheet(self, sheetName: str) -> None:
        '''
        Switch to a sheet.

        name: the name of the sheet
        '''

        self.activeWB.Sheets(sheetName).Activate()

    def GetSheetsState(self, add_example_data2feedback=False) -> str:
        '''
        Get the state of all sheets.

        Param:
        @ add_example_data2feedback: add example data of each column to the description
        
        Returns:
        str -- the state of the sheets.
        '''
        states = []
        for ws in self.activeWB.Worksheets:
            if ws.Range('A1').Value is None:
                cell_state = "Sheet \"{}\" {} has no content".format(ws.Name, '(active)' if ws.Name == self.activeWS.Name else '')
            else:
                NumberOfRows = ws.UsedRange.Rows.Count
                NumberOfColumns = ws.UsedRange.Columns.Count
                headers = ws.Range('A1', ws.Cells(1,NumberOfColumns)).Value
                if isinstance(headers, tuple):
                    headers = headers[0]
                else:
                    headers = [headers]
                headers = {get_column_letter(i): header for i, header in enumerate(headers, start=1)}
                
                if add_example_data2feedback:
                    col_example_data_str = []
                    # Iterate through the columns
                    for column_index, (col_letter, header) in enumerate(headers.items(), start=1):  # Assuming you want the first four columns
                        column = ws.Columns(column_index)
                        cell_range = column.Range("A1:A5")  # Adjust the range as needed

                        # Get values from the cells in the range
                        values = [str(cell.Text) for cell in cell_range]
                        data_type = ws.Cells(2, column_index).NumberFormat.replace('G/通用格式', 'General')
                        
                        col_example_data_str.append("{}: {{".format(col_letter) + f"Header: {values[0]}, Data type: {data_type}, 4 sample rows: [{', '.join(values[1:])}]" + '}')
                    
                    cell_state = 'Sheet \"{}\" has {} columns ({}) and {} rows (1 header row and {} data rows)'.format(ws.Name, NumberOfColumns, ', '.join(col_example_data_str), NumberOfRows, NumberOfRows-1)
                else:                    
                    cell_state = 'Sheet \"{}\" has {} columns (Headers are {}) and {} rows (1 header row and {} data rows)'.format(ws.Name, NumberOfColumns, ', '.join(["{}: \"{}\"".format(col_letter, header) for col_letter, header in headers.items()]), NumberOfRows, NumberOfRows-1)
            
            # Add chart descriptions
            # Iterate through the shapes to find chart objects
            chart_names = []
            for shape in ws.Shapes:
                if shape.HasChart:
                    chart = shape.Chart
                    chart_name = chart.Name
                    chart_names.append(chart_name[chart_name.find(' ')+1:])
                    
            chartNameString = ' and this sheet has the charts whose names are "{}"'.format('", "'.join(chart_names)) if len(chart_names) > 0 else ''
            
            # Iterate through the pivot tables and print their names
            pt_names = []
            for pivot_table in ws.PivotTables():
                pt_names.append(pivot_table.Name)

            if len(pt_names) > 0:
                ptNameString = ' the pivot tables whose names are "{}"'.format('", "'.join(pt_names))
                if len(chart_names) == 0:
                    ptNameString = ' and this sheet has' + ptNameString
                else:
                    ptNameString = ' and' + ptNameString
            else:
                ptNameString = ''
            
            states.append("{}{}{}.".format(cell_state, chartNameString, ptNameString))
                                                                                                                       
        return "Sheet state: " + ' '.join(states)

    # def Filter

if __name__ == '__main__':
    bot = xwBackend()
    #bot.Write('A1', [['1']])
    bot.CreatePivotTable('Sheet1!A1:B3', 'Sheet1', 'PivotTable1', ['A'], [], [], ['B'], 'avg')
    bot.AutoFill('Sheet1!G2', 'Sheet1!G2:G36')
    bot.CopyPaste('A1', 'A2')
