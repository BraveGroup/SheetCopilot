# This script is used to paraphrase SU adapted instructions so that the paraphrases are more like what human users would say when useing Excel
# The original instructions are too unnatural and lenghty.
import openpyxl, openai, os, re, tiktoken, tqdm, time, json
from collections import defaultdict
import requests
os.environ["http_proxy"] = "http://127.0.0.1:7890"
os.environ["https_proxy"] = "http://127.0.0.1:7890"

from chatgpt_wrapper import ChatGPT
from utils import num_tokens_from_string, generate_state

encoding = tiktoken.encoding_for_model('gpt-3.5-turbo')

from utils import num_tokens_from_string, generate_state, ask

openai.api_key = ""


PROMPT = """You have been tasked with paraphrasing a set of instructions for Excel tasks.

Requirements:
1. Paraphrase each given instruction so that it is more like what a non-expert Excel user would say while retaining the original intention and order of actions in the instruction.
2. Do not mention any Excel built-in functions in the paraphrases. If the original instruction mentions the names of specific Excel built-in functions, you should replace these functions with spoken language. For example, "Use CONCATENATE to combine cells" mentions "CONCATENATE" and should be rephrased as "concatenate cells". "Use conditional formatting to change the background color of cells" mentions "conditional formatting" and should be rephrased as "Set cell background color if ...". "Write a formula to copy data" should be rephrased as "Copy data".
3. Do not refer to existing columns by their indices (e.g., column A is not desired); instead, mention columns by their headers. For instance, 'In column A' should be rephrased as 'In the Year column'.
4. Do not add column references in brackets. For instance, "subtracting Total Expenses (Column B) from Revenue (Column A)" should be rephrased as "Subtract Total Expenses from Revenue".
5. When inserting new columns, the column indices must be given to avoid ambiguity. Besides, each new column must be given a header, so you need to keep the column headers mentioned in the original instructions.
6. Use domain-specific knowledge to diversify the expression of the generated instructions and do not use the same expression repeatedly.

I will give you some examples first:

Original instructions:
1. Convert the "Year" column values (Column A) from the format of yyyy.00 to yyyy.
2. In a new column with the header "Combined data", use the CONCATENATE function to combine the data from columns A through Z in each row, including the headers, and then autofill the formula to the last row of data.
3. Use advanced filters in "Sheet1" to display rows with sales data for a specific month (assuming weeks starting from the 1st of each month). For example, filter to show rows with weeks "Week 1" to "Week 4" for the first month.
4. Create a line chart on a new sheet named "Sheet2" with weeks (Column A) on the x-axis and sales (Column B) as the y-axis.
5. Apply conditional formatting to Column H (Net Profit) based on Columns F (Operating Profit) and G (Tax Expense). If the cell value in Column G is not 0 and Column F is greater than 0, display the value in Column H in red.
6. In a new sheet named "Sheet3", use the VLOOKUP function to match each row in "Year" (Column A) in "Sheet1" with the corresponding value in Column B ("Net Sales") from "Sheet1".
7. In a new column (Column G) with the header "Tax Calculation", use a formula to calculate the tax by multiplying the corresponding "Subtotal" (Column D) by 0.1.

Think about the flaws in the original instructions before paraphrasing:
1.
Think: The original instruction refers to columns by indices in brackets, which is undesired.
Paraphrase: Convert the "Year" column format from yyyy.00 to yyyy.

2.
Think: The original instruction mentions Excel built-in functions CONCATENATE, which is undesired.
Paraphrase: Concatenate the data from columns A through Z for all rows and write the results in a new column named "Combined data".

3.
Think: The original instruction mentions an Excel built-in operation (i.e., advanced filters), which is undesired.
Paraphrase: Display only the rows with weeks from Week 1 to Week 4.

4.
Think: The original instruction refers to columns by indices in brackets, which is undesired.
Paraphrase: Create a line chart on a new sheet with the Week column on the x-axis and the Sales column as the y-axis.

5.
Think: The original instruction mentions Excel built-in functions (i.e., conditional formatting) and refers to columns by indices, which is undesired.
Paraphrase: If the cell value in Tax Expense Column is not 0 and that in Operating Profit Column > 0, display the cell text in the Net Profit column in red.

6.
Think: The original instruction mentions Excel built-in functions (i.e., VLOOKUP) and refers to columns by indices in brackets, which is undesired.
Paraphrase: Match cells in the Year column and return the corresponding values in the Net Sales Column. Put the results in a new sheet.

7.
Think: The original instruction refers to columns by indices in brackets, which is undesired.
Paraphrase: Calculate the tax by multiplying the "Subtotal" column by 0.1 in column G named "Tax Calculation".

Now it's your turn. Please follow the requirements to paraphrase the original instructions according to the given workbook descriptions.
"""

def main():
    base_dir = r"D:\Github\ActionTransformer\Excel_data\example_sheets_part1"
    data_path = os.path.join(base_dir, "SU_adaptation_v2")
    output_dir = os.path.join(base_dir, "SU_adaptation_v2_rephrase")
    os.makedirs(output_dir, exist_ok=True)

    task_classes = ['Finance'] # os.listdir(data_path)

    batch_size = 6
    
    gpt_mode = ['wrapper', 'api', 'proxy'][2]
    debug = False
    bot = None
    if gpt_mode == 'wrapper':
        # config = Config()
        bot = ChatGPT()
    
    for task_class in task_classes:
        log_file = os.path.join(output_dir, f'{task_class}_ckpt.json')
        while True:
            try:
                # Loading instructions to be rephrased
                task_instruction_xls = openpyxl.load_workbook(os.path.join(data_path, '{}_adaptation.xlsx'.format(task_class)))
                instruction_sheet = task_instruction_xls.worksheets[0]
                print('Processing {} ({} instructions to be rephrased  Batch size {})'.format(task_class, instruction_sheet.max_row - 1, batch_size))

                new_task_xls = os.path.join(output_dir, '{}.xlsx'.format(task_class))

                if os.path.exists(new_task_xls):
                    new_task = openpyxl.load_workbook(new_task_xls)
                    sheet = new_task.worksheets[0]
                          
                    if os.path.exists(log_file):
                        # read response log
                        with open(log_file, 'r') as f:
                            log = json.load(f)
                        
                        ckpt, response_log = log['ckpt'], log['response']

                        print("Loading checkpoint ({} samples have been rephrased)".format(len(ckpt)))

                else:
                    new_task = openpyxl.Workbook()
                    sheet = new_task.worksheets[0]

                    if os.path.exists(log_file):
                        log_file.replace(".json", "_1.json")
                    # simplify the above lone
                    sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'], sheet['F1'], sheet['G1'] = 'Sheet Name', 'Context', 'Instructions', 'Source', 'Categories', 'Atomic ations', 'Seed task'
                    ckpt, response_log = [], []

                row_to_write = sheet.max_row + 1

                task_dict = defaultdict(list)
                for row_id in range(2, instruction_sheet.max_row + 1):
                    task_dict[instruction_sheet[f'A{row_id}'].value].append(row_id)
                

                for task_id, (task_sheet_name, row_ids) in enumerate(task_dict.items()):
                    if debug and task_id > 3: break

                    task_book = openpyxl.load_workbook(os.path.join(r"D:\Github\ActionTransformer\Excel_data\example_sheets_part1", task_class, f'{task_sheet_name}.xlsx'))
                    task_context = instruction_sheet[f'B{row_ids[0]}'].value.strip()
                    task_state = "\nGiven an Excel workbook:\n{}\n\n".format(task_context + ' ' + generate_state(task_book))

                    with tqdm.tqdm(total=len(row_ids), desc="Processing Task {}: {}".format(task_id, task_sheet_name)) as pbar:
                    # Collect samples
                        while True:
                            tasks = []

                            while len(row_ids) > 0 and len(tasks) < batch_size:
                                row_id = row_ids.pop(0)
                                if row_id in ckpt:
                                    pbar.update(1)
                                    continue

                                ckpt.append(row_id)               
                                tasks.append({'row_id': row_id, 'task_name': task_sheet_name, 'context': task_context, 'instruction': instruction_sheet[f'C{row_id}'].value.strip(), 'categories': instruction_sheet[f'E{row_id}'].value, 'atomic_actions': instruction_sheet[f'F{row_id}'].value, 'seed_task':  instruction_sheet[f'G{row_id}'].value})

                            if len(tasks) == 0:
                                task_book.close()
                                break
                            
                            instruc_to_rephrase = '\n\n'.join(['{}. {}'.format(j+1, x['instruction']) for j, x in enumerate(tasks)])

                            input_text = PROMPT + task_state + "Original instructions:\n" + instruc_to_rephrase + f"\n\nThink about the flaws in the original instructions before paraphrasing:"

                            response_text = ask(input_text, gpt_mode, bot)

                            response_log.append({'row_ids': [x['row_id'] for x in tasks], 'task_names': [x['task_name'] for x in tasks], 'input_text': input_text, 'response_text': response_text})

                            matches = [x.end() for x in re.finditer("Paraphrase:", response_text)]

                            if len(matches) == 0:
                                # Extracting the sentence from response_text after each No. in the form of "[0-9.]"
                                matches = [x.end() for x in re.finditer(r"\d+\.\s", response_text)]
                            
                            assert len(matches) == len(tasks), "GPT output incorrect!"

                            for sample, match_id in zip(tasks, matches):
                                para = response_text[match_id:response_text.find('\n', match_id)].strip()

                                sheet[f"A{row_to_write}"], sheet[f"B{row_to_write}"], sheet[f"C{row_to_write}"] = sample['task_name'], sample['context'], para

                                sheet[f"E{row_to_write}"] = sample['categories']; sheet[f"F{row_to_write}"], sheet[f'G{row_to_write}'] = sample['atomic_actions'], sample['seed_task']

                                row_to_write += 1
                            
                            new_task.save(new_task_xls)

                            # save log
                            with open(log_file, 'w') as f:
                                json.dump({'ckpt': ckpt, 'response': response_log}, f, indent=2)

                            pbar.update(len(tasks))

                            if debug: break
                break
            except Exception as e:
                pass

    print('All {} tasks have been self-instructed and saved to {}. Quitting...'.format(new_task_xls, instruction_sheet.max_row))

main()