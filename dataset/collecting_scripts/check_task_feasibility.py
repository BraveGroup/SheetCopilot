import openpyxl, openai, os, re, tqdm, time, json, requests
from collections import defaultdict
os.environ["http_proxy"] = "http://127.0.0.1:7890"
os.environ["https_proxy"] = "http://127.0.0.1:7890"

from chatgpt_wrapper import ChatGPT

from utils import num_tokens_from_string, generate_state, ask
# from chatgpt_wrapper.core.config import Config

openai.api_key = ""

PROMPT = """I will give you a batch of Excel task instructions that are utilized to evaluate the spreadsheet manipulation capabilities of large language models. Please check all instructions according to the criteria and the descriptions of the given workbooks.

Requirements:

1. If an instruction fulfills all of the following four criteria strictly at the same time, it is valid.
A. Realism: Verify that the instructions are representative of real-world Excel problems encountered by expert users. Avoid including contrived or overly complex tasks that would rarely be encountered in practice.
B. Relevance: Verify that the instructions are relevant to the context of the given workbooks. Reject instructions that do not relate to the content of the workbooks or are not applicable in the context of Excel.
C. Clarity: Verify that the instructions are clear and easy to understand. They should be free from grammatical errors and ambiguities.
D. Completeness: Verify that the instructions can be completed using the provided workbook data and Excel features. If additional data or functionality is needed to accomplish a task, reject the instruction.

2. Use your domain-specific knowledge to reason and make decisions. For example, it is unlikely to calculate monthly or yearly averages of some physical values because this task is impractical.

I will give you an example first:

Given an Excel workbook:
My workbook records many invoices made on different dates. Sheet "Sheet1" has 7 columns (Headers are A: "No.", B: "Date", C: "Salesman", D: "Product", E: "Price", F: "Units", G: "Sales") and 19 rows (including the header row). The cells in the "No." column range from 10500.00 to 10505.00. The cells in the "Date" column can be "2011-05-28 00:00:00", "2011-05-25 00:00:00", "2011-05-27 00:00:00". The cells in the "Salesman" column can be "Joe", "Chin", "Moe". The cells in the "Product" column can be "Quad", "Majestic", "Bellen", "Carlota", "Alpine". The cells in the "Price" column range from 22.00 to 32.00. The cells in the "Units" column range from 4.00 to 25.00. The cells in the "Sales" column range from 128.00 to 750.00.

Instructions to check:
1. Find the sales value corresponding to each No. in a new column called "Invoice Lookup".
2. Compare the names in the Salesman column to find the closest match for each name. Put the results in a new column named "Salesman Matched".
3. Create a sheet named "Sheet2" to summarize the total sales for each sales representative in Sheet1.
4. Prepend leading zeros to each number in the "No." column so that they have a fixed length of 6 digits. Append the new results to the corresponding product names in the "Product" column, and put the results in a new column named "Padded No.".
5. Find the corresponding date for each No. in a new column named "Lookup Date" in Sheet1.
6. Find and display the row values in Sheet1 based on the "No." column. Put the results in a new sheet named "Sheet2".
7. Round the values in the "Sales" column to two decimal places in a new column named "Rounded Sales", and display the results with trailing zeros.
8. Merge cells A1 through C1 with cells A2 through D2.
9. Add hyperlinks to each cell in the "No." column that link to its corresponding file.

Check results (Give brief reasons in the comments and tell me if the instruction is valid):
1. Realism: Yes, Relevance: Yes, Clarity: Yes, Completeness: Yes. Comment: The instruction fulfills the 4 criteria, so it is valid. 
2. Realism: No, Relevance: No, Clarity: No, Completeness: No. Comment: This instruction is unrealistic and unclear and does not seem to be relevant to the context of the given workbook, so it is invalid.
3. Realism: Yes, Relevance: Yes, Clarity: Yes, Completeness: Yes. Comment: The instruction fulfills the 4 criteria, so it is valid. 
4. Realism: No, Relevance: Yes, Clarity: Yes, Completeness: Yes. Comment: The instruction appends numbers to product names, which is not a realistic requirement, so it is invalid.
5. Realism: Yes, Relevance: Yes, Clarity: Yes, Completeness: Yes. Comment: The instruction fulfills the 4 criteria, so it is valid.
6. Realism: Yes, Relevance: Yes, Clarity: No, Completeness: No. Comment: The instruction does not specify what values to display, so it is invalid.
7. Realism: Yes, Relevance: Yes, Clarity: Yes, Completeness: Yes. Comment: The instruction fulfills the 4 criteria, so it is valid.
8. Realism: No, Relevance: No, Clarity: Yes, Completeness: Yes. Comment: The instruction merges cells, which destroys the original data and is meaningless in the context of the workbook, so it is invalid.
9. Realism: Yes, Relevance: Yes, Clarity: Yes, Completeness: No. Comment: The instruction does not refer to specific corresponding files, which is incomplete, so it is invalid.

Now it's your turn.

Given an Excel workbook:
{}

Instructions to check:
{}

Check results (Give brief reasons in the comments and tell me if the instruction is valid):
"""

def main():
    data_path = "./"
    output_dir = os.path.join(data_path, "SU_adaptation_check")
    batch_size = 14

    gpt_mode = ['wrapper', 'api', 'proxy'][2]
    debug = False
    bot = None
    if gpt_mode == 'wrapper':
        # config = Config()
        bot = ChatGPT()
    
    log_file = os.path.join(output_dir, f'check_ckpt.json')
    output_file = os.path.join(output_dir, f'check.xlsx')
    while True:
        try:
            print('Processing...')
            task_file = os.path.join(data_path, "SU_rephrase", f"adaptation.xlsx")
            task_instruction_xls = openpyxl.load_workbook(task_file)
            instruction_sheet = task_instruction_xls.worksheets[0]

            if os.path.exists(log_file):
                    # read response log
                    with open(log_file, 'r') as f:
                        log = json.load(f)
                    
                    ckpt, response_log = log['ckpt'], log['response']

                    print("Loading checkpoint ({} samples have been rephrased)".format(len(ckpt)))

            else:
                instruction_sheet['H1'], instruction_sheet['I1'] = 'Validity', 'Reason'

                if os.path.exists(log_file):
                    log_file.replace(".json", "_1.json")
                
                ckpt, response_log = [], []

            task_dict = defaultdict(list)
            for row_id in range(2, instruction_sheet.max_row + 1):
                task_dict[instruction_sheet[f'A{row_id}'].value].append(row_id)

            for task_id, (task_sheet_name, row_ids) in enumerate(task_dict.items()):
                task_book = openpyxl.load_workbook(os.path.join(data_path, f'{task_sheet_name}.xlsx'))
                task_context = instruction_sheet[f'B{row_ids[0]}'].value.strip()
                task_state = task_context + ' ' + generate_state(task_book)

                with tqdm.tqdm(total=len(row_ids), desc="Processing Task {}: {}".format(task_id + 1, task_sheet_name)) as pbar:
                # Collect samples
                    while True:
                        tasks = []

                        while len(row_ids) > 0 and len(tasks) < batch_size:
                            row_id = row_ids.pop(0)
                            if row_id in ckpt:
                                pbar.update(1)
                                continue

                            ckpt.append(row_id)               
                            tasks.append({'row_id': row_id, 'task_name': task_sheet_name, 'context': task_context, 'instruction': instruction_sheet[f'C{row_id}'].value.strip(), 'categories': instruction_sheet[f'E{row_id}'].value, 'atomic_actions': instruction_sheet[f'F{row_id}'].value})

                        if len(tasks) == 0:
                            task_book.close()
                            break
                        
                        instruc_to_check = '\n'.join(['{}. {}'.format(j+1, x['instruction']) for j, x in enumerate(tasks)])

                        input_text = PROMPT.format(task_state, instruc_to_check)

                        response_text = ask(input_text, gpt_mode, bot)

                        response_log.append({'row_ids': [x['row_id'] for x in tasks], 'task_names': [x['task_name'] for x in tasks], 'input_text': input_text, 'response_text': response_text})

                        matches = [x.end() for x in re.finditer("Comment:", response_text)]

                        assert len(matches) == len(tasks), "GPT output incorrect!"

                        for sample, match_id in zip(tasks, matches):
                            check_result = response_text[match_id:response_text.find('\n', match_id)].strip()
                            label = 'Valid' if 'invalid' not in check_result.lower() else 'Invalid'
                            instruction_sheet[f"I{sample['row_id']}"] = check_result[:check_result.rfind(',')]
                            instruction_sheet[f"H{sample['row_id']}"] = label
                        
                        task_instruction_xls.save(output_file)

                        # save log
                        with open(log_file, 'w') as f:
                            json.dump({'ckpt': ckpt, 'response': response_log}, f, indent=2)
                        
                        pbar.update(len(tasks))

                        if debug: break
            break
        except Exception as e:
            raise e

    print('All {} tasks have been checked. Saving to {}'.format(instruction_sheet.max_row, output_file))

main()