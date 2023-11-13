import openpyxl, openai, os, re, tqdm, time, json, requests
from collections import defaultdict
os.environ["http_proxy"] = "http://127.0.0.1:7890"
os.environ["https_proxy"] = "http://127.0.0.1:7890"

try:
    from chatgpt_wrapper import ChatGPT
except:
    pass

from utils import num_tokens_from_string, generate_state, ask

openai.api_key = ""


RAW_PROMPT = """As an Excel expert, you have been assigned the responsibility of adapting a set of task instructions for specific Excel workbooks. These instructions will be utilized to evaluate the Excel manipulation capabilities of large language models.

Requirements:
1. First, identify individual atomic actions used in the original instructions, then develop new instructions incorporating these actions.
2. Use the detailed descriptions of the provided workbooks to modify the original instructions so that they become compatible with the workbooks. Specifically, you must change the manipulated objects (ranges, sheets, rows, and columns) in the original instructions. You must also change the way you use atomic actions. For instance, if the original instruction sets the data type as accounting, you can change it to other types in the adaptation.
3. Use standard range references, such as 'Sheet2!A1:C9', 'A2:E16', 'A:H', column C, or row 16.
4. Use different phrasing (e.g., various sentence structures and noun/verb forms) to create diverse instructions.
5. Apply domain-specific knowledge to diversify the generated instructions. For instance, use financial knowledge to calculate various metrics, demonstrate trends in product sales from different perspectives, visualize data using various types of charts, and so on.
6. (Important!) The generated instructions must describe realistic tasks and involve the normal use of atomic actions according to the workbook descriptions.
7. In every new instruction, new tables and Pivot tables must be created in a new worksheet and start from A1. Only one new sheet is allowed to be created. Besides, the headers of new columns/rows and the names of new sheets must be set.
8. The instructions after adaptation should look like what a non-expert Excel user would say and should not mention any specific funtions or operations built in Excel.

Here are the atomic actions you can identify within the six categories:
{}

Restrictions for the atomic action parameters:
Chart type can only be (stacked/clustered) column chart, (stacked/clustered) bar chart, (3D) pie chart, line chart (with smooth lines), (stacked) area chart, or scatter chart.
Cell value type can only be date, text, time, currency, percentage, number, or general.

I will give you an example first:

Given an Excel workbook:
The sheet 'Sheet1' records the employee working hours of a coffee shop. It has 8 columns (Headers are: A: "location", B: "name", C: "date", D: "hours", E: "ot hours", F: "base pay", G: "ot pay", H: "total pay") and 11 rows (including the header row). The cells in the "location" column can be "capitol hill", "queen anne". The cells in the "name" column can be "Aaron", "Bob", "Blanca", "Frank".

The original instructions to be adapted:
1. I'd like to know if there's a formula (or combination of formulas) to sum Column B ONLY if it equals Column A?
2. Column A contains multiple due dates and cell B1 is 10/31/2022. Append a special character to the cells in column A depending on whether the due date is less or more than the date in B1. If neither applies, leave the cell unchanged.
3. Create a Pivot Table to separate dates by each month similar to the quarterly function. This will show all dates in the last year under the column label. Choose the monthly option under data filters in the columns label under data filters.
4. Freeze A1:B10 so that no matter how I scroll vertically or horizontally this range is always frozen.
5. I have five groups of data each occupying two columns. I'd like to have them plotted all on one bar chart but with the series separated (i.e., not clustered) so that several column charts stick together sharing the same y-axis.


Adaptations compatible with the given workbook (Show the categories involved in the generated instruction and list the atomic actions following the category label):
1. Instruction: In a new column with header "Total pay each location", use a formula to sum the "total pay" (Column H) for each employee ONLY if their "location" (Column A) is "capitol hill". - Categories (atomic actions): A (Update cell value); F (Math functions)

2. Instruction: Create a formula in a new column (Column I) with header "Marked Due Dates" to check if the dates in column C are less or more than 10/31/2022. If the date is less, append a '-' to the cell in the same row in column C; if the date is more, append a '+'; otherwise, leave the cell unchanged. - Categories (atomic actions): A (Update cell value, Autofill); F (Logical functions, Text functions)

3. Instruction: Create a Pivot Table in a new sheet named "Sheet2" based on the data in 'Sheet1' and then summarize sum of hours for each location in this Pivot Table. - Categories (atomic actions): A (Create sheet); E (Create Pivot Table)

4. Instruction: Freeze the range A1:H1 so that the headers remain visible when scrolling vertically or horizontally. - Categories (atomic actions): C (Freeze panes)

5. Instruction: Create a Pivot Table in a new sheet named "Sheet2" to sum the hours for all dates and then plot a line chart to show the trend of hours changing with dates. - Categories (atomic actions): A (Create sheet); E (Create Pivot Table); D (Create chart)

Now it's your turn.
"""

cate_dict = dict(zip(['A','B','C','D','E','F'], ['Entry and manipulation', 'Management', 'Formatting', 'Charts', 'Pivot Table', 'Formula']))

def read_atomic_action(file):
    atomic_action_wb = openpyxl.load_workbook(file)

    cat_info_lst = []
    for k, v in cate_dict.items():
        seed_task_ws = atomic_action_wb[v]
        cat_info_lst.append('{}. {}: {}'.format(k, v, ', '.join(seed_task_ws[f'A{i}'].value for i in range(2, seed_task_ws.max_row + 1))))

    cat_info = '\n'.join(cat_info_lst)

    return cat_info

def main():
    data_path = "../"
    atomic_action_file = os.path.dirname(__file__) + '/atomic_actions.xlsx'

    batch_size = 10

    seed_task_file = os.path.dirname(__file__) + '/seed_tasks.xlsx'
    seed_task_wb = openpyxl.load_workbook(seed_task_file)
    seed_task_ws = seed_task_wb.worksheets[0]
    
    cat_info = read_atomic_action(atomic_action_file)

    PROMPT = RAW_PROMPT.format(cat_info)

    use_wrapper = True
    debug = False

    gpt_mode = ['wrapper', 'api', 'proxy'][2]
    debug = False
    bot = None
    if gpt_mode == 'wrapper':
        # config = Config()
        bot = ChatGPT()
    
    output_dir = os.path.join(data_path, 'SU_adaptation')

    retry_id = 0
    ckpt_file = os.path.join(output_dir, f'adaptation_ckpt.txt')
    adapt_stats_file = os.path.join(output_dir, f'adapt_stats.json')
    response_log = os.path.join(output_dir, f'adapation_response_log.txt')

    while True:
        print('Processing ...' + "Retry: {}".format(retry_id))
        retry_id += 1

        new_task_xls = os.path.join(output_dir, f'adaptation.xlsx')

        last_row_id_this_class, last_row_id_seed_task = 1, 1

        if os.path.exists(new_task_xls):
            new_task = openpyxl.load_workbook(new_task_xls)
            sheet = new_task.worksheets[0]

            if os.path.exists(ckpt_file):
                with open(ckpt_file, 'r') as f:
                    last_row_id_this_class, last_row_id_seed_task = map(int, f.read().strip().split())

                    if last_row_id_seed_task == seed_task_ws.max_row:
                        last_row_id_this_class += 1
                        last_row_id_seed_task = 1 

                    print("Resume from the {} th row of task insturction file and {} th row of the seed task file".format(last_row_id_this_class + 1, last_row_id_seed_task + 1))
            
            with open(adapt_stats_file, 'r') as f:
                adapt_stats = defaultdict(dict)
                for k,v in json.load(f).items():
                    adapt_stats[k] = v
        else:
            new_task = openpyxl.Workbook()
            sheet = new_task.worksheets[0]
            sheet['A1'] = 'Sheet Name'; sheet['B1'] = 'Context'; sheet['C1'] = 'Instructions'; sheet['D1'] = 'Source'; sheet['E1'] = 'Categories'; sheet['F1'] = 'Atomic ations'
            adapt_stats = defaultdict(dict)
        
        row_to_write = sheet.max_row + 1
        
        # We need the spreadsheet names and contexts to adapt seed tasks. This info is stored in task_instructions.xlsx
        task_instruction_xls = openpyxl.load_workbook(os.path.join('../task_sheets', 'task_instructions.xlsx'))
        instruction_sheet = task_instruction_xls.worksheets[0]

        if seed_task_ws.max_row == last_row_id_seed_task:break

        num_batches = (seed_task_ws.max_row - last_row_id_seed_task) // batch_size
        
        # Adapt all
        for i in range(last_row_id_this_class + 1, instruction_sheet.max_row + 1):
            if debug and i == 4: break

            # Skip duplicate task
            if instruction_sheet[f'A{i}'].value.strip() == instruction_sheet[f'A{i-1}'].value.strip(): continue

            task_sheet_name = instruction_sheet[f'A{i}'].value.strip()
            task_context = instruction_sheet[f'C{i}'].value.strip()
            task_book = openpyxl.load_workbook(os.path.join(data_path, f'task_sheets/{task_sheet_name}.xlsx'))

            task_state = "\nGiven an Excel workbook:\n{}\n".format(task_context + ' ' + generate_state(task_book))
            
            for batch_id in tqdm.trange(num_batches + 1):
                if debug and batch_id == 1: break

                batch_start = last_row_id_seed_task + 1
                batch_end = min(seed_task_ws.max_row + 1, batch_start + batch_size)
                seed_task_batch = [[seed_task_ws[f'A{j}'].value, seed_task_ws[f'B{j}'].value] for j in range(batch_start, batch_end)]

                seed_task_text = "\nThe original instructions to be adapted:\n{}\n".format('\n'.join("{}. {}".format(j+1, seed_task_batch[j][1]) for j in range(len(seed_task_batch))))

                input_text = PROMPT + task_state + seed_task_text + "\nAdaptations compatible with the given workbook (Show the categories involved in the generated instruction and list the atomic actions following the category label):\n"
                print("Input token len:", num_tokens_from_string(input_text))

                # with open('temp.txt', 'w') as f:
                #     f.write(input_text)
                
                response_text = ask(input_text, gpt_mode, bot)

                # with open('./gpt_output.txt', 'r') as f:
                #     response_text = f.read()

                matches = re.finditer("Instruction:", response_text)

                for j, match_id in enumerate(matches):
                    instruc_cat_acts = response_text[match_id.end():response_text.find('\n', match_id.end())].strip()

                    for NA_keyword in ['adapted', 'applicable']:
                        if NA_keyword in instruc_cat_acts.lower():
                            adapt_stats[seed_task_batch[j][0]][task_sheet_name] = 0
                            break
                    else:
                        adapt_stats[seed_task_batch[j][0]][task_sheet_name] = 1
                    
                    if adapt_stats[seed_task_batch[j][0]][task_sheet_name] == 0: continue

                    adapt_stats[seed_task_batch[j][0]][task_sheet_name] = 1

                    for keyword in ["- Categories (atomic actions):", "- Category (atomic actions):", "- Categories (atomic action):", "- Category (atomic action):"]:
                        if keyword in instruc_cat_acts:
                            instruc, cat_acts = instruc_cat_acts.split(keyword)
                            break
                    
                    # Extract task categories and corresponding atomic actions
                    cats = set(); acts = []

                    pattern = r'([A-Z]) \(([^)]+)\)'
                    matches = re.findall(pattern, cat_acts)
                    
                    for match in matches:
                        cats_text = match[0].strip()
                        cats.add(cats_text if cats_text in cate_dict.values() else cate_dict[cats_text])
                        acts.extend([a.strip() for a in match[1].split(',')])
                    
                    sheet[f"A{row_to_write}"], sheet[f"B{row_to_write}"], sheet[f"C{row_to_write}"] = task_sheet_name, instruction_sheet[f'C{i}'].value, instruc.strip()
                    if sum(')' in x for x in acts) > 0: raise Exception
                    sheet[f"E{row_to_write}"] = ', '.join(list(cats)); sheet[f"F{row_to_write}"] = ', '.join(acts)

                    row_to_write += 1
                
                last_row_id_seed_task += len(seed_task_batch)

                # Save checkpoint
                new_task.save(new_task_xls)

                # Save adapt statistics as json file
                with open(adapt_stats_file, 'w') as f:
                    json.dump(adapt_stats, f)
                
                with open(ckpt_file, 'w') as f:
                    f.write('{} {}'.format(last_row_id_this_class, last_row_id_seed_task))
                    print('Saving checkpoint. Last processed task: {}. Last processed seed task {}'.format( last_row_id_this_class, last_row_id_seed_task))
                
                with open(response_log, 'a') as f:
                    f.write('Conversation:\n{}\n{}\n\n'.format(input_text, response_text))

                if last_row_id_seed_task == seed_task_ws.max_row: break
            
            
            last_row_id_this_class += 1

            task_book.close()

            if last_row_id_seed_task == seed_task_ws.max_row:
                print('All seed tasks have been adapted for {}/{} instruction.'.format(i - 1, instruction_sheet.max_row - 1))
                if last_row_id_this_class == instruction_sheet.max_row:
                    print('All {} tasks have been self-istructed. Processing next class...'.format(instruction_sheet.max_row))
                    break

            # reset eed task pointer
            last_row_id_seed_task = 1

        new_task.close()
        task_instruction_xls.close()

    seed_task_wb.close()

main()